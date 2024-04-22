import os
import re
from collections import OrderedDict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
import sys
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
if __name__ == "__main__":
    # Accept the Excel file path as an argument
    excel_file_path = sys.argv[1]

    # Read the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    excel_name = excel_file_path.split('/')[-1]
    excel_name = os.path.splitext(os.path.basename(excel_file_path))[0].replace('_'," ")+" Assist"
    print(excel_name)


if 'matrix' in excel_name.lower():
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Category":
                cell.value = "Sub_Category"  
sheet['A1'].value = None
cell_to_unmerge = 'A1'
if sheet.merged_cells:
    merged_cells_copy = sheet.merged_cells.ranges.copy()  
    for merged_cell in merged_cells_copy:
        if cell_to_unmerge in merged_cell:
            sheet.unmerge_cells(str(merged_cell))

if 'beam' in excel_name.lower():
    insert_column_index = 1  
    sheet.insert_cols(insert_column_index)
    for row in sheet.iter_rows():
        for cell in row:
            if (cell.value == 'glare_rate_oncoming_400m [%]' or cell.value == 'mat_glare_rate_oncoming_400m [%]') and cell.row > 1:
                sheet.cell(row=cell.row, column=insert_column_index).value = 'Glare rates'       
            elif cell.value == 'spatial_driver_visible_oncoming_all' and cell.row > 1:
                sheet.cell(row=cell.row, column=insert_column_index).value = 'False Positive Rates'
                sheet.cell(row=cell.row - 1, column=insert_column_index).value = 'Category'
         

def Remove_unwanted_rows(sheet):
    sum_of_weights_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value == "Sum of weights"]
    threshold_low_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value == "Threshold low"]
    perception_version_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value == "Perception_Version"]
    for sum_of_weights_cell in sum_of_weights_cells:
        empty_cells_rows = [row[0].row for row in sheet.iter_rows(min_row=sum_of_weights_cell.row + 1, min_col=sum_of_weights_cell.column, max_col=sum_of_weights_cell.column) if not row[0].value]
        for row_index in sorted(empty_cells_rows, reverse=True):
            sheet.delete_rows(row_index)
        sheet.delete_rows(sum_of_weights_cell.row)
    for threshold_low_cell in threshold_low_cells:
        sheet.delete_rows(threshold_low_cell.row)
    for perception_version_cell in perception_version_cells:
        row_index = perception_version_cell.row
        sheet.delete_rows(row_index)  
        sheet.delete_rows(row_index)  

replacements = {
    "Value": "Actual","Threshold high": "Target","Number non-corrupted HiL re-runs":"No. Of Clips evaluated","glare_rate_oncoming_400m [%]":"<=400m","glare_rate_oncoming_all [%]":"All","glare_rate_preceding_300m [%]":"<=300","glare_rate_preceding_all [%]":"All","glare_rate_position_lights [%]":"All",
    "spatial_driver_visible_oncoming_all":"oncoming_all","spatial_driver_visible_rate_preceding_all":"preceding_all","spatial_marginal_rate_oncoming_all":"oncoming_all","spatial_marginal_rate_preceding_all":"preceding_all","mat_glare_rate_oncoming_400m [%]":"All (>400m)","mat_glare_rate_oncoming_all [%]":"<400m",
    "mat_glare_rate_position_lights [%]":"All (>300m)","mat_glare_rate_preceding_300m [%]":"<300m","mat_glare_rate_preceding_all [%]":"All"

}

def add_new_string(sheet, row, column, new_string):
    previous_cell = sheet.cell(row=row, column=column - 1)
    previous_cell.value = new_string

def add_new_category(sheet, KPI_Names):
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell_value = cell.value
            if isinstance(cell_value, str):  # Check if cell_value is a string
                for kpi in KPI_Names:
                    if kpi in cell_value:  # Check if the KPI name is present in the cell value
                        if 'driver_visible' in kpi:
                            add_new_string(sheet, row, column, 'Driver Visible')
                        elif 'marginal_rate' in kpi:
                            add_new_string(sheet, row, column, 'Marginal Rate')
                        elif 'position_lights' in kpi:
                            add_new_string(sheet, row, column, 'Position_lights')
                        elif 'preceding' in kpi:
                            add_new_string(sheet, row, column, 'Preceding')
                        elif 'oncoming' in kpi:
                            add_new_string(sheet, row, column, 'Oncoming')
                        break 

def replace_old_values(sheet,replacements):
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell_value = cell.value
            old_value = cell.value
            if old_value in replacements:
                cell.value = replacements[old_value]

def merge_cells(sheet, Type_values, Cells_to_merge, new_values_merge):
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell_value = cell.value
            merge_start_row = row + 1
            merge_end_column = column + 1
            while merge_start_row <= sheet.max_row:
                next_cell = sheet.cell(row=merge_start_row, column=column)
                if next_cell.value == cell_value:
                    merge_start_row += 1
                    continue
                elif next_cell.value or next_cell.value == 0 or next_cell.value in Cells_to_merge or merge_start_row in new_indices:
                    break
                merge_start_row += 1
            if cell_value in Type_values:
                while merge_end_column <= sheet.max_column:
                    next_cell = sheet.cell(row=row, column=merge_end_column)
                    if next_cell.value in Type_values or next_cell.coordinate in sheet.merged_cells:
                        break
                    merge_end_column += 1
                if merge_end_column - column > 1:
                    sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=merge_end_column - 1)
            if cell_value in Cells_to_merge or cell_value == 'Matrix Beam Assistant' or cell_value in new_values_merge:
                if merge_start_row - row > 1:
                    if cell_value in new_values_merge:
                        sheet.merge_cells(start_row=row, start_column=column, end_row=merge_start_row - 1, end_column=column + 1)
                    else:
                        sheet.merge_cells(start_row=row, start_column=column, end_row=merge_start_row - 1, end_column=column)

def bold_matching_cells_in_excel(sheet, values_to_bold):
    bold_font = Font(bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in values_to_bold or cell.value=='Sub_Category' or cell.value=='Country':
                cell.font = bold_font
 
def add_values_above_first_occurrence(sheet, value_to_find, value):
    found = False
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value == value_to_find:
                found = True
                row_index = cell.row
                col_index = cell.column
                for i, new_value in enumerate(value):
                    sheet.cell(row=row_index - i - 1, column=col_index).value = new_value
                return True  
value_to_find = 'Value'
new_values = ['Weather', 'Light_condition', 'Road_Type']
new_values1=['Road_Type','Country']

def find_category(sheet, search_values):
    data = {value: [] for value in search_values}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in search_values:
                for below_cell in sheet.iter_rows(min_row=cell.row + 1, min_col=cell.column, max_col=cell.column):
                    for value_cell in below_cell:
                        if isinstance(value_cell.value, str) and value_cell.value.strip():  
                            data[cell.value].append(value_cell.value)
    return data

search_values = ["KPI_Name", "Category", "Sub_Category"]

def find_Type_values(sheet, values):
    KPI_Names = OrderedDict()  
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in values:
                for right_column in range(cell.column + 1, sheet.max_column + 1):  
                    right_cell = sheet.cell(row=cell.row, column=right_column)
                    if isinstance(right_cell.value, str) and right_cell.value.strip():
                        KPI_Names[right_cell.value] = None  
 
    return list(KPI_Names.keys())

def add_total(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ["Day", "Overcast", "City"]:
                previous_cell = sheet.cell(row=cell.row, column=cell.column - 1)
                previous_cell_value = previous_cell.value.strip()
                if previous_cell_value == "":
                    previous_cell.value = str(previous_cell.value) + "Total"
                elif previous_cell.value is not None:
                    previous_cell = sheet.cell(row=cell.row, column=cell.column - 2)
                    previous_cell.value = str(previous_cell.value) + "Total"

def extract_specific_values(KPI_Names):
    specific_values = []
    for kpi in KPI_Names:
        try:
            if 'driver_visible' in kpi:
                specific_values.append('Driver Visible')
            elif 'marginal_rate' in kpi:
                specific_values.append('Marginal Rate')
            elif 'position_lights' in kpi:
                specific_values.append('Position_lights')
            elif 'preceding' in kpi:
                specific_values.append('Preceding')
            elif 'oncoming' in kpi:
                specific_values.append('Oncoming')
        except Exception as e:
            print(f"Error processing KPI '{kpi}': {e}")
    return specific_values

def apply_border_to_cells(sheet):
    total_cell = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Total":
                total_cell = cell
                break
        if total_cell:
            break
    if total_cell:
        max_row = sheet.max_row
        max_column = sheet.max_column
        border_medium = Border(left=Side(style='medium'), right=Side(style='medium'),
                               top=Side(style='medium'), bottom=Side(style='medium'))
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet.iter_rows(min_row=total_cell.row, min_col=total_cell.column, max_row=max_row, max_col=max_column):
            for cell in row:
                if cell.row not in new_indices:
                    cell.border = border_medium
        value_actual_cell = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "Actual":
                    value_actual_cell = cell
                    break
            if value_actual_cell:
                break

        if value_actual_cell:
            for row in sheet.iter_rows(min_row=value_actual_cell.row, min_col=1, max_row=max_row, max_col=value_actual_cell.column):
                for cell in row:
                    if cell.row not in new_indices:
                        cell.border = border_thin
def find_rows_range(sheet, value):
    rows_range = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                start_row = cell.row
                next_row = start_row + 1
                while next_row <= sheet.max_row:
                    if sheet.cell(row=next_row, column=cell.column).value in (None, ''):
                        next_row += 1
                    else:
                        break
                rows_range.extend(range(start_row, next_row))  
    return rows_range
def cut_and_insert_rows_above(sheet, rows, target_row):
    for idx, row_num in enumerate(rows, 1):
        for col_num, cell in enumerate(sheet[row_num], 1):
            sheet.cell(row=target_row - 1 - idx, column=col_num, value=cell.value)
        sheet.delete_rows(row_num)
rows_range = find_rows_range(sheet, "in_city")
target_row = next((cell.row for row in sheet.iter_rows() for cell in row if cell.value == 'high beam assistant(rates)'), None)
if rows_range and target_row:
    cut_and_insert_rows_above(sheet, rows_range[::-1], target_row)

def process_excel_sheet(sheet):
    # Find the Country Column
    country_column = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Country":
                country_column = cell.column
                break
        if country_column:
            break
    new_column_letter = get_column_letter(country_column + 1)
    sheet.insert_cols(country_column + 1)
    grand_total_column = None
    grand_total_column = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Grand Total":
                grand_total_column = cell.column
                cell_above = sheet.cell(row=cell.row - 1, column=cell.column)
                cell_above.value = "Total"
                cell.value = None  # Clear the Grand Total cell
                break
        if grand_total_column:
            break
    for row in sheet.iter_rows(min_row=2, min_col=grand_total_column, max_col=grand_total_column):
        for cell in row:
            new_cell = sheet.cell(row=cell.row, column=country_column + 1)
            new_cell.value = cell.value
            # Copy cell fill color
            if cell.fill is not None:
                new_cell.fill = PatternFill(start_color=cell.fill.start_color.rgb,
                                            end_color=cell.fill.end_color.rgb,
                                            fill_type=cell.fill.fill_type)
            # Copy cell borders
            if cell.border is not None:
                new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)

def find_new_indices(sheet):
    indices = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ["Matrix Beam Assistant", "High Beam Assistant"]:
                indices.append(cell.row)

    new_indices = []  # Initialize new_indices list to keep track of adjusted indices
    offset = 0  # Initialize offset
    for index in sorted(indices):
        sheet.insert_rows(index + offset)
        new_indices.append(index + offset)  # Add adjusted index to new_indices
        offset += 1  # Increase offset since we inserted a row
    return new_indices

if 'country' in excel_name.lower():
    insert_column_index = 1  
    sheet.insert_cols(insert_column_index)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'mat_glare_rate_oncoming_400m [%]' and cell.row > 1:
                    sheet.cell(row=cell.row, column=insert_column_index).value = 'Glare rates'       
            elif cell.value == 'spatial_driver_visible_oncoming_all' and cell.row > 1:
                    sheet.cell(row=cell.row, column=insert_column_index).value = 'False Positive Rates'
    insert_column_index = 1  
    sheet.insert_cols(insert_column_index)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'city_entry' and cell.row > 1:
                    sheet.cell(row=cell.row, column=insert_column_index).value = 'City Detection' 
            elif cell.value == 'high beam assistant(rates)' and cell.row > 1:
                    sheet.cell(row=cell.row, column=insert_column_index).value = 'High Beam Assistant'
            elif cell.value == 'False Positive Rates' and cell.row > 1:
                    sheet.cell(row=cell.row, column=insert_column_index).value = 'Matrix Beam Assistant'   

new_values_merge=["City Detection","High Beam Assistant"]
new_colums=["Glare rates","False Positive Rates","Total"]
cell=sheet['A1']
cell.value=excel_name
font = Font(size=14, name='Arial')  
cell.font = font
Remove_unwanted_rows(sheet)
if "city" in excel_name.lower() or "beam" in excel_name.lower():
    add_values_above_first_occurrence(sheet, value_to_find, new_values)
    add_total(sheet)
if 'country' in excel_name.lower():
    add_values_above_first_occurrence(sheet, value_to_find, new_values1)
    process_excel_sheet(sheet)
data_dict = find_category(sheet, search_values)
KPI_Names = data_dict["KPI_Name"]
Type_search_values=set(new_values+new_values1)
Type_values=find_Type_values(sheet, Type_search_values)
add_new_category(sheet, KPI_Names)
data_dict = find_category(sheet, search_values)
category = data_dict["Category"]
sub_category = data_dict["Sub_Category"]
Cells_to_merge=category+KPI_Names+new_colums+sub_category+new_values_merge
new_indices=find_new_indices(sheet)
merge_cells(sheet, Type_values, Cells_to_merge, new_values_merge)
values_to_bold=Type_values+category+new_values+search_values+search_values+new_colums+sub_category
bold_matching_cells_in_excel(sheet, values_to_bold)
replace_old_values(sheet,replacements)
values_to_border=category+KPI_Names
apply_border_to_cells(sheet)

if "City" in excel_name:
    KPI_Names_updated = [re.sub(r'\[.*?\]', '[1/km]', kpi_name) for kpi_name in KPI_Names]
    for row_num in range(1, sheet.max_row + 1):
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value in KPI_Names:
                    cell.value = KPI_Names_updated[KPI_Names.index(cell.value)]
for row in sheet.iter_rows():
    for cell in row:
        if cell.coordinate != 'A1':
            cell.alignment = Alignment(horizontal='center', vertical='center')

for row in sheet.iter_rows():
    for cell in row:
        content_length = len(str(cell.value))
        width = (content_length + 2) * 1.2
        column = cell.column
        sheet.column_dimensions[get_column_letter(column)].width = max(width, sheet.column_dimensions[get_column_letter(column)].width)
modified_file_path = excel_file_path.replace('.xlsx', '_Report.xlsx')
workbook.save(modified_file_path)
print("Report Generated Succesfully")